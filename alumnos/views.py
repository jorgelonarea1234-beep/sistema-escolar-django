from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Count, Avg, Q
from .models import Alumno, Materia, Calificacion
from .forms import AlumnoForm, MateriaForm, CalificacionForm, RegularizacionForm
import openpyxl
from openpyxl import load_workbook
from .models import Carrera, Materia
from django.shortcuts import render, redirect, get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth.models import User
import random
import string
from django.db.models import Max
from django.shortcuts import redirect
import json
from django.http import JsonResponse
from .models import Calificacion, Alumno, Materia
from .forms import MaestroForm
from .models import Maestro
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.models import Group
from django.http import JsonResponse
import json
from .models import Horario
from .models import ConfiguracionParcial
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.shortcuts import get_object_or_404, redirect, render
from .models import ConfiguracionRegularizacion
from .constants import SEMESTRES



def guardar_parciales(request):
    import json
    from django.http import JsonResponse
    from .models import Calificacion, Alumno, Materia, ConfiguracionParcial

    data = json.loads(request.body)

    alumno = Alumno.objects.get(id=data['alumno_id'])
    materia = Materia.objects.get(id=data['materia_id'])

    def guardar(parcial, valor):

        habilitado = ConfiguracionParcial.objects.filter(
            parcial=parcial,
            habilitado=True
        ).exists()

        if not habilitado:
            return

        if valor == "" or valor is None:
            Calificacion.objects.filter(
                alumno=alumno,
                materia=materia,
                parcial=parcial
            ).delete()
            return

        try:
            valor = int(valor)
        except:
            raise ValueError("La calificación debe ser un número válido.")

        if valor < 0 or valor > 100:
            raise ValueError("La calificación debe estar entre 0 y 100.")

        Calificacion.objects.update_or_create(
            alumno=alumno,
            materia=materia,
            parcial=parcial,
            defaults={
                'calificacion': valor
            }
        )

    try:
        guardar(1, data.get('p1'))
        guardar(2, data.get('p2'))
        guardar(3, data.get('p3'))

        return JsonResponse({'status': 'ok'})

    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'mensaje': str(e)
        })


def eliminar_parciales(request):
    import json
    from django.http import JsonResponse
    from .models import Calificacion, Alumno, Materia

    data = json.loads(request.body)

    alumno = Alumno.objects.get(id=data['alumno_id'])
    materia = Materia.objects.get(id=data['materia_id'])

    Calificacion.objects.filter(alumno=alumno, materia=materia).delete()

    return JsonResponse({'status': 'ok'})


def agregar_horario(request):

    if request.method == 'POST':
        data = json.loads(request.body)

        materia = Materia.objects.get(id=data['materia_id'])

        Horario.objects.create(
            materia=materia,
            dia=data['dia'],
            hora_inicio=data['hora_inicio'],
            hora_fin=data['hora_fin']
        )

        return JsonResponse({'status': 'ok'})


def eliminar_horario(request):

    if request.method == 'POST':
        data = json.loads(request.body)

        horario = Horario.objects.get(id=data['id'])
        horario.delete()

        return JsonResponse({'status': 'ok'})


@login_required
def horario_alumno(request):

    alumno = Alumno.objects.get(user=request.user)
    materias = alumno.materias.prefetch_related('horarios')

    dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
    bloques = [7, 9, 11, 13]

    return render(request, 'alumnos/horario.html', {
        'materias': materias,
        'dias': dias,
        'bloques': bloques
    })


@login_required
def inscripcion_materias(request):

    alumno = Alumno.objects.get(user=request.user)
    materias = Materia.objects.all()

    semestres = range(1, 9)

    return render(request, 'alumnos/inscripcion_materias.html', {
        'materias': materias,
        'alumno': alumno,
        'semestres': semestres
    })



def inscribir_materia(request):

    if request.method == 'POST':

        alumno = Alumno.objects.get(user=request.user)

        # 🔥 AQUÍ VA LO QUE PREGUNTASTE
        data = json.loads(request.body)
        materia_id = data.get('materia_id')

        materia = Materia.objects.get(id=materia_id)
        materias_actuales = alumno.materias.all()

        # 🔥 VALIDACIÓN
        if hay_conflicto(materia, materias_actuales):
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Hay choque de horario en materias seleccionadas'
            })

        alumno.materias.add(materia)

        return JsonResponse({'status': 'ok'})

    return JsonResponse({'status': 'error'}, status=400)




def hay_conflicto(nueva_materia, materias_alumno):

    for materia_existente in materias_alumno:

        for h1 in nueva_materia.horarios.all():
            for h2 in materia_existente.horarios.all():

                # mismo día
                if h1.dia != h2.dia:
                    continue

                # choque de horario
                if (
                    h1.hora_inicio < h2.hora_fin and
                    h1.hora_fin > h2.hora_inicio
                ):
                    return True

    return False



def eliminar_materia_ajax(request):

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            materia_id = data.get('id')

            if not materia_id:
                return JsonResponse({
                    'success': False,
                    'message': 'ID no recibido'
                })

            materia = Materia.objects.get(id=materia_id)
            materia.delete()

            return JsonResponse({'success': True})

        except Materia.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Materia no existe'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    return JsonResponse({'success': False})

def eliminar_carrera_ajax(request):

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            carrera = Carrera.objects.get(id=data.get('id'))
            carrera.delete()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False})

@login_required
def editar_carrera(request, id):

    carrera = get_object_or_404(Carrera, id=id)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        print("NUEVO NOMBRE:", nombre)  # 🔥 DEBUG

        carrera.nombre = nombre
        carrera.save()

        return JsonResponse({'success': True})

    return render(request, 'alumnos/editar_carrera_modal.html', {
        'carrera': carrera
    })


def es_admin(user):
    return user.is_superuser

@user_passes_test(es_admin)
def eliminar_carrera(request, id):

    carrera = get_object_or_404(Carrera, id=id)
    carrera.delete()

    return redirect('lista_carreras')

@permission_required('alumnos.delete_materia', raise_exception=True)
def eliminar_materia(request, id):

    materia = get_object_or_404(Materia, id=id)
    materia.delete()

    return redirect('lista_materias')



@login_required
def editar_materia(request, id):

    materia = get_object_or_404(Materia, id=id)

    if request.method == 'POST':
        form = MateriaForm(request.POST, instance=materia)

        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})

    else:
        form = MateriaForm(instance=materia)

    return render(request, 'alumnos/editar_materia_modal.html', {
        'form': form,
        'materia': materia
    })




def login_view(request):

    print("ANTES:", request.user.is_authenticated)

    if request.method == 'POST':
        correo = request.POST.get('correo')
        password = request.POST.get('password')

        user = authenticate(request, username=correo, password=password)

        print("USER:", user)

        if user is not None:
            login(request, user)

            print("DESPUÉS:", request.user.is_authenticated)  # 🔥

            return redirect('/alumnos/')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'registration/login.html')



def logout_view(request):
    logout(request)
    return redirect('login')



def crear_maestro_ajax(request):
    import json
    from django.http import JsonResponse
    from .models import Maestro

    data = json.loads(request.body)

    Maestro.objects.create(nombre=data['nombre'])

    return JsonResponse({'status': 'ok'})




def editar_maestro_ajax(request):
    import json
    from django.http import JsonResponse
    from .models import Maestro

    data = json.loads(request.body)

    maestro = Maestro.objects.get(id=data['id'])
    maestro.nombre = data['nombre']
    maestro.save()

    return JsonResponse({'status': 'ok'})



from django.http import JsonResponse
import json
from .models import Maestro

def eliminar_maestro_ajax(request):

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            print("DATA RECIBIDA:", data)  # 🔥 DEBUG

            maestro_id = data.get('id')

            if not maestro_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID no recibido'
                }, status=400)

            maestro = Maestro.objects.get(id=maestro_id)
            maestro.delete()

            return JsonResponse({'status': 'ok'})

        except Maestro.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Maestro no existe'
            }, status=404)

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

    return JsonResponse({'status': 'error'}, status=400)





def lista_maestros(request):

    maestros = Maestro.objects.all()

    return render(request, 'alumnos/lista_maestros.html', {
        'maestros': maestros
    })



def crear_maestro(request):

    if request.method == 'POST':

        nombre = request.POST.get('nombre')
        correo = request.POST.get('correo')
        password = request.POST.get('password')

        # 🔥 VALIDAR CORREO
        if User.objects.filter(username=correo).exists():
            messages.error(request, "El correo ya está registrado")
            return render(request, 'alumnos/crear_maestro.html')

        # 🔥 CREAR USUARIO (CORREGIDO)
        user = User.objects.create_user(
            username=correo,
            email=correo,  # 👈🔥 IMPORTANTE
            password=password
        )

        Maestro.objects.create(user=user, nombre=nombre)

        messages.success(request, "Maestro creado correctamente")

        return redirect('lista_maestros')

    return render(request, 'alumnos/crear_maestro.html')





def crear_calificacion_ajax(request):
    import json
    from django.http import JsonResponse
    from .models import Calificacion, Alumno, Materia

    try:
        data = json.loads(request.body)

        print("DATA RECIBIDA:", data)

        alumno_id = data.get('alumno_id')
        materia_id = data.get('materia_id')
        calificacion_val = data.get('calificacion')

        if not alumno_id or not materia_id or not calificacion_val:
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Datos incompletos'
            })

        try:
            calificacion = float(calificacion_val)
        except:
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Calificación inválida'
            })

        if calificacion < 1 or calificacion > 100:
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Calificación inválida (1-100)'
            })

        alumno = Alumno.objects.get(id=alumno_id)
        materia = Materia.objects.get(id=materia_id)

        calificacion_obj, created = Calificacion.objects.update_or_create(
            alumno=alumno,
            materia=materia,
            defaults={'calificacion': calificacion}
        )
        
        return JsonResponse({
            'status': 'ok',
            'id': calificacion_obj.id  # 🔥 ESTE ES EL ID REAL
        })

    except Exception as e:
        print("🔥 ERROR REAL:", str(e))
        return JsonResponse({
            'status': 'error',
            'mensaje': str(e)
        })


@login_required
@permission_required('alumnos.change_calificacion', raise_exception=True)
def lista_regularizaciones(request):

    config = ConfiguracionRegularizacion.objects.first()
    if not config or not config.habilitado:
        messages.error(request, "Las regularizaciones aún no están habilitadas.")
        return redirect('lista_alumnos')

    calificaciones = Calificacion.objects.filter(
        calificacion__lt=70
    ).select_related(
        'alumno',
        'materia'
    ).order_by(
        'alumno__nombre',
        'materia__nombre',
        'parcial'
    )

    return render(
        request,
        'alumnos/regularizaciones/lista.html',
        {
            'calificaciones': calificaciones
        }
    )



@login_required
@permission_required('alumnos.change_calificacion', raise_exception=True)
def capturar_regularizacion(request, pk):

    config = ConfiguracionRegularizacion.objects.first()
    if not config or not config.habilitado:
        messages.error(request, "Las regularizaciones aún no están habilitadas.")
        return redirect('lista_alumnos')

    calificacion = get_object_or_404(
        Calificacion,
        pk=pk
    )

    if request.method == 'POST':

        form = RegularizacionForm(
            request.POST,
            instance=calificacion
        )

        if form.is_valid():
            form.save()
            return redirect('lista_regularizaciones')

    else:

        form = RegularizacionForm(
            instance=calificacion
        )

    return render(
        request,
        'alumnos/regularizaciones/form.html',
        {
            'form': form,
            'calificacion': calificacion
        }
    )


def obtener_kardex(alumno):

    materias = alumno.materias.all().order_by('semestre', 'nombre')

    kardex = {}
    promedio_general_lista = []
    aprobadas = 0
    reprobadas = 0

    for materia in materias:

        semestre = materia.semestre

        if semestre not in kardex:
            kardex[semestre] = {
                'nombre': SEMESTRES.get(
                    semestre,
                    f"Semestre {semestre}"
                ),
                'materias': [],
                'promedio': 0,
                'aprobadas': 0,
                'reprobadas': 0,
                'cursadas': 0
            }

        calificaciones = Calificacion.objects.filter(
            alumno=alumno,
            materia=materia
        )

        notas = []

        for c in calificaciones:
            notas.append(c.calificacion_final())

        if notas:
            promedio_materia = round(sum(notas) / len(notas), 2)
        else:
            promedio_materia = None

        if promedio_materia is not None:
            promedio_general_lista.append(promedio_materia)

            if promedio_materia >= 70:
                estado = 'Aprobada'
                aprobadas += 1
                kardex[semestre]['aprobadas'] += 1
            else:
                estado = 'Reprobada'
                reprobadas += 1
                kardex[semestre]['reprobadas'] += 1
        else:
            estado = 'Sin calificar'

        tipo = 'Ordinario'

        for c in calificaciones:
            if c.calificacion < 70 and c.regularizacion is not None:
                tipo = 'Regularización'

        kardex[semestre]['materias'].append({
            'nombre': materia.nombre,
            'final': promedio_materia,
            'estado': estado,
            'tipo': tipo
        })

        kardex[semestre]['cursadas'] += 1

    for semestre, datos in kardex.items():

        finales = [
            m['final']
            for m in datos['materias']
            if m['final'] is not None
        ]

        if finales:
            datos['promedio'] = round(sum(finales) / len(finales), 2)
        else:
            datos['promedio'] = None

    promedio_general = round(
        sum(promedio_general_lista) / len(promedio_general_lista),
        2
    ) if promedio_general_lista else None

    return {
        'kardex': kardex,
        'promedio_general': promedio_general,
        'aprobadas': aprobadas,
        'reprobadas': reprobadas,
        'materias_cursadas': aprobadas + reprobadas
    }



@login_required
def kardex_alumno(request, id):

    alumno = Alumno.objects.get(id=id)

    if hasattr(request.user, 'alumno') and request.user.alumno.id != alumno.id:
        return redirect('lista_alumnos')

    datos_kardex = obtener_kardex(alumno)

    return render(
        request,
        'alumnos/kardex.html',
        {
            'alumno': alumno,
            **datos_kardex
        }
    )


def guardar_calificacion_ajax(request):
    import json
    from django.http import JsonResponse
    from .models import Calificacion, Materia, Alumno

    data = json.loads(request.body)

    calificacion_valor = int(data['calificacion'])

    if calificacion_valor < 1 or calificacion_valor > 100:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Calificación inválida (1-100)'
        })

    materia = Materia.objects.get(id=data['materia_id'])
    alumno = Alumno.objects.get(id=data['alumno_id'])
    parcial = int(data['parcial'])  # 🔥 NUEVO

    Calificacion.objects.update_or_create(
        alumno=alumno,
        materia=materia,
        parcial=parcial,  # 🔥 CLAVE
        defaults={'calificacion': calificacion_valor}
    )

    return JsonResponse({'status': 'ok'})




def editar_calificacion_ajax(request):
    import json
    from django.http import JsonResponse
    from .models import Calificacion

    data = json.loads(request.body)

    cal = Calificacion.objects.get(id=data['id'])

    # 🔥 VALIDACIÓN DE SEGURIDAD
    if hasattr(request.user, 'maestro'):
        if cal.materia.maestro != request.user.maestro:
            return JsonResponse({'error': 'No autorizado'}, status=403)

    # 🔥 ACTUALIZAR
    cal.calificacion = data['calificacion']
    cal.save()

    return JsonResponse({'status': 'ok'})


from django.http import JsonResponse
from .models import Calificacion
import json

def eliminar_calificacion_ajax(request):

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            cal_id = data.get('id')

            if not cal_id:
                return JsonResponse({'success': False, 'error': 'ID no enviado'})

            cal = Calificacion.objects.filter(id=cal_id).first()

            if not cal:
                return JsonResponse({'success': False, 'error': 'No existe la calificación'})

            cal.delete()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def dashboard(request):

    user = request.user

    # 👑 ADMIN
    if user.is_superuser or user.groups.filter(name='admin').exists():

        context = {
            'tipo': 'admin',
            'total_alumnos': Alumno.objects.count(),
            'total_maestros': Maestro.objects.count(),
            'total_materias': Materia.objects.count(),
        }

    # 👨‍🏫 MAESTRO
    elif user.groups.filter(name='maestro').exists():

        maestro = Maestro.objects.get(user=user)

        materias = Materia.objects.filter(maestro=maestro)

        alumnos = Alumno.objects.filter(
            materias__maestro=maestro
        ).distinct()

        context = {
            'tipo': 'maestro',
            'materias': materias.count(),
            'alumnos': alumnos.count(),
        }

    # 👨‍🎓 ALUMNO
    elif user.groups.filter(name='alumno').exists():

        alumno = Alumno.objects.get(user=user)

        context = {
            'tipo': 'alumno',
            'nombre': alumno.nombre,
            'matricula': alumno.matricula,
            'carrera': alumno.carrera.nombre,
            'materias': alumno.materias.count(),
        }

    else:
        context = {'tipo': 'otro'}


    context.update({
    'es_admin': user.is_superuser or user.groups.filter(name='admin').exists(),
    'es_maestro': user.groups.filter(name='maestro').exists(),
    'es_alumno': user.groups.filter(name='alumno').exists(),
})    

    return render(request, 'alumnos/dashboard.html', context)


# ===============================
# 👨‍🎓 ALUMNOS
# ===============================
@login_required
def lista_carreras(request):
    carreras = Carrera.objects.all()

    return render(request, 'alumnos/lista_carreras.html', {
        'carreras': carreras
    })


from django.contrib.auth.decorators import login_required
from .models import Alumno, Maestro, Materia


@login_required
def lista_alumnos(request):
    user = request.user

    # 👑 ADMIN
    if user.is_superuser or user.groups.filter(name='admin').exists():
        alumnos = Alumno.objects.all()

    # 👨‍🏫 MAESTRO
    elif user.groups.filter(name='maestro').exists():

        try:
            maestro = Maestro.objects.get(user=user)

            alumnos = Alumno.objects.filter(
                materias__maestro=maestro
            ).distinct()

        except Maestro.DoesNotExist:
            alumnos = Alumno.objects.none()

    # 👨‍🎓 ALUMNO 👇 AQUÍ VA LO TUYO
    elif user.groups.filter(name='alumno').exists():

        try:
            alumno = Alumno.objects.get(user=user)
            alumnos = Alumno.objects.filter(id=alumno.id)

        except Alumno.DoesNotExist:
            alumnos = Alumno.objects.none()

    # 🔒 OTROS
    else:
        alumnos = Alumno.objects.none()

    return render(request, 'alumnos/lista_alumnos.html', {
        'alumnos': alumnos,
        'es_admin': user.is_superuser or user.groups.filter(name='admin').exists(),
        'es_maestro': user.groups.filter(name='maestro').exists(),
        'es_alumno': user.groups.filter(name='alumno').exists(),
    })




def crear_carrera(request):

    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        Carrera.objects.create(nombre=nombre.upper())

        return redirect('lista_carreras')

    return render(request, 'alumnos/crear_carrera.html')

def asignar_materias_carrera(request, carrera_id):

    carrera = get_object_or_404(Carrera, id=carrera_id)
    materias = Materia.objects.all()

    if request.method == 'POST':
        materias_ids = request.POST.getlist('materias')

        carrera.materia_set.set(materias_ids)  # 🔥 relación

        return redirect('lista_carreras')  # o donde quieras

    return render(request, 'alumnos/asignar_materias.html', {
        'carrera': carrera,
        'materias': materias,
        'materias_asignadas': carrera.materia_set.all()
    })


def generar_matricula(self):
    ultimo = Alumno.objects.exclude(matricula__isnull=True)\
                           .exclude(matricula__exact='')\
                           .order_by('-id')\
                           .first()
    
    if ultimo and '-' in ultimo.matricula:
        try:
            numero = int(ultimo.matricula.split('-')[-1]) + 1
        except:
            numero = 1
    else:
        numero = 1

    return f"ALU-{numero:04d}"



@login_required
@permission_required('alumnos.add_alumno', raise_exception=True)
def crear_alumno(request):

    carreras = Carrera.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        correo = request.POST.get('correo')
        password = request.POST.get('password')
        carrera_id = request.POST.get('carrera')

        # 🔥 validar usuario
        if User.objects.filter(username=correo).exists():
            messages.error(request, "El correo ya está registrado")
            return redirect('crear_alumno')

        # 🔥 crear usuario
        user = User.objects.create_user(
            username=correo,
            email=correo,
            password=password
        )

        # 🔥 obtener carrera
        carrera = Carrera.objects.get(id=carrera_id)

        # 🔥 crear alumno
        Alumno.objects.create(
            nombre=nombre,
            correo=correo,
            carrera=carrera,
            user=user
        )

        messages.success(request, "Alumno creado correctamente")
        return redirect('lista_alumnos')

    return render(request, 'alumnos/crear_alumno.html', {
        'carreras': carreras
    })






@login_required
@permission_required('alumnos.change_alumno', raise_exception=True)
def editar_alumno(request, id):

    alumno = Alumno.objects.get(id=id)

    if request.method == 'POST':
        form = AlumnoForm(request.POST, instance=alumno)

        if form.is_valid():
            form.save()

            messages.success(request, "Alumno actualizado correctamente")  # 🔥 AQUÍ

            return redirect('lista_alumnos')

    else:
        form = AlumnoForm(instance=alumno)

    return render(request, 'alumnos/editar_alumno.html', {
        'form': form
    })




@login_required
@permission_required('alumnos.delete_alumno', raise_exception=True)
def eliminar_alumno(request, id):

    alumno = Alumno.objects.get(id=id)
    alumno.delete()

    messages.success(request, 'Alumno eliminado correctamente')
    return redirect('lista_alumnos')


# ===============================
# 📚 MATERIAS
# ===============================

@login_required
def lista_materias(request):

    materias = Materia.objects.all()
    return render(request, 'alumnos/lista_materias.html', {'materias': materias})




@login_required
@permission_required('alumnos.add_materia', raise_exception=True)
def crear_materia(request):

    if request.method == 'POST':
        form = MateriaForm(request.POST)

        if form.is_valid():
            materia = form.save(commit=False)

            if hasattr(request.user, 'maestro'):
                materia.maestro = request.user.maestro

            materia.save()
            form.save_m2m()

            messages.success(request, 'Materia creada correctamente')
            return redirect('lista_materias')

    else:
        form = MateriaForm()  # 🔥 ESTA LÍNEA FALTABA

    return render(request, 'alumnos/crear_materia.html', {'form': form})

# ===============================
# 📊 DETALLE + CALIFICACIONES
# ===============================
@login_required
def detalle_alumno(request, id):

    alumno = Alumno.objects.get(id=id)

    materias = alumno.materias.all()

    calificaciones = Calificacion.objects.filter(alumno=alumno)

    parciales_habilitados = {
       c.parcial: c.habilitado
       for c in ConfiguracionParcial.objects.all()
    }

    mapa = {}

    for c in calificaciones:

        if c.materia_id not in mapa:
            mapa[c.materia_id] = {
                'p1': '',
                'r1': '',
                'p2': '',
                'r2': '',
                'p3': '',
                'r3': '',
                'promedio': ''
            }

        if c.parcial == 1:
            mapa[c.materia_id]['p1'] = c.calificacion
            mapa[c.materia_id]['r1'] = c.regularizacion or ''

        elif c.parcial == 2:
            mapa[c.materia_id]['p2'] = c.calificacion
            mapa[c.materia_id]['r2'] = c.regularizacion or ''

        elif c.parcial == 3:
            mapa[c.materia_id]['p3'] = c.calificacion
            mapa[c.materia_id]['r3'] = c.regularizacion or ''

    for key in mapa:
        notas = []

        p1 = mapa[key]['p1']
        r1 = mapa[key]['r1']

        p2 = mapa[key]['p2']
        r2 = mapa[key]['r2']

        p3 = mapa[key]['p3']
        r3 = mapa[key]['r3']

        if p1 != '':
            if p1 < 70 and r1 != '':
                notas.append(r1)
            else:
                notas.append(p1)

        if p2 != '':
            if p2 < 70 and r2 != '':
                notas.append(r2)
            else:
                notas.append(p2)

        if p3 != '':
            if p3 < 70 and r3 != '':
                notas.append(r3)
            else:
                notas.append(p3)

        if notas:
            promedio = sum(notas) / len(notas)
            mapa[key]['promedio'] = round(promedio, 1)
        else:
            mapa[key]['promedio'] = ''

    return render(
        request,
        'alumnos/detalle_alumno.html',
        {
            'alumno': alumno,
            'materias': materias,
            'mapa': mapa,
            'parciales_habilitados': parciales_habilitados
        }
    )  


# ===============================
# ➕ CALIFICACIONES
# ===============================

def crear_calificacion(request, alumno_id):

    alumno = Alumno.objects.get(id=alumno_id)

    # 🔥 AQUÍ VA
    print("MATERIAS DEL ALUMNO:", alumno.materias.all())


    form = CalificacionForm(request.POST or None)
    form.fields['materia'].queryset = alumno.materias.all()

    if request.method == 'POST':
        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.alumno = alumno
            calificacion.save()
            return redirect('detalle_alumno', id=alumno.id)

    return render(request, 'alumnos/crear_calificacion.html', {
        'form': form,
        'alumno': alumno
    })


@login_required
@permission_required('alumnos.change_calificacion', raise_exception=True)
def editar_calificacion(request, id):

    calificacion = Calificacion.objects.get(id=id)

    if request.method == 'POST':
        form = CalificacionForm(request.POST, instance=calificacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Calificación actualizada correctamente")
            return redirect('detalle_alumno', id=calificacion.alumno.id)
    else:
        form = CalificacionForm(instance=calificacion)

    return render(request, 'alumnos/crear_calificacion.html', {'form': form})


@login_required
@permission_required('alumnos.delete_calificacion', raise_exception=True)
def eliminar_calificacion(request, id):

    calificacion = Calificacion.objects.get(id=id)
    alumno_id = calificacion.alumno.id
    calificacion.delete()

    messages.success(request, "Calificación eliminada correctamente")

    return redirect('detalle_alumno', id=alumno_id)


# ===============================
# 📄 PDF
# ===============================

@login_required
def generar_pdf(request, id):

    alumno = Alumno.objects.get(id=id)
    materias = alumno.materias.all()
    calificaciones = Calificacion.objects.filter(alumno=alumno)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="boleta_{alumno.matricula}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    elementos = []

    elementos.append(Paragraph("SISTEMA ESCOLAR", styles['Title']))
    elementos.append(Paragraph(f"<b>Boleta de calificaciones</b>", styles['Heading2']))
    elementos.append(Spacer(1, 12))

    elementos.append(Paragraph(f"<b>Alumno:</b> {alumno.nombre}", styles['Normal']))
    elementos.append(Paragraph(f"<b>Matrícula:</b> {alumno.matricula}", styles['Normal']))
    elementos.append(Paragraph(f"<b>Correo:</b> {alumno.correo}", styles['Normal']))
    elementos.append(Paragraph(f"<b>Carrera:</b> {alumno.carrera}", styles['Normal']))
    elementos.append(Spacer(1, 20))

    data = [['Materia', 'P1', 'R1', 'P2', 'R2', 'P3', 'R3', 'Promedio']]

    for materia in materias:
        p1 = calificaciones.filter(materia=materia, parcial=1).first()
        p2 = calificaciones.filter(materia=materia, parcial=2).first()
        p3 = calificaciones.filter(materia=materia, parcial=3).first()
    
        v1 = p1.calificacion if p1 else ''
        r1 = p1.regularizacion if p1 and p1.regularizacion is not None else ''
    
        v2 = p2.calificacion if p2 else ''
        r2 = p2.regularizacion if p2 and p2.regularizacion is not None else ''
    
        v3 = p3.calificacion if p3 else ''
        r3 = p3.regularizacion if p3 and p3.regularizacion is not None else ''
    
        valores = []
    
        if p1:
            valores.append(p1.calificacion_final())
    
        if p2:
            valores.append(p2.calificacion_final())
    
        if p3:
            valores.append(p3.calificacion_final())
    
        promedio = round(sum(valores) / len(valores), 2) if valores else ''
    
        data.append([
            materia.nombre,
            str(v1),
            str(r1),
            str(v2),
            str(r2),
            str(v3),
            str(r3),
            str(promedio)
        ])

    table = Table(data, colWidths=[170, 45, 45, 45, 45, 45, 45, 65])    

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
    ]))

    elementos.append(table)
    elementos.append(Spacer(1, 20))

    doc.build(elementos)

    return response

# ===============================
# 📥📤 EXCEL
# ===============================

@login_required
def exportar_excel(request):

    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = 'Alumnos'

    hoja['A1'] = 'Matricula'
    hoja['B1'] = 'Nombre'
    hoja['C1'] = 'Correo'
    hoja['D1'] = 'Carrera'

    alumnos = Alumno.objects.all()

    fila = 2
    for alumno in alumnos:
        hoja[f'A{fila}'] = alumno.matricula
        hoja[f'B{fila}'] = alumno.nombre
        hoja[f'C{fila}'] = alumno.correo
        hoja[f'D{fila}'] = alumno.carrera
        fila += 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="alumnos.xlsx"'

    workbook.save(response)
    return response


@login_required
def importar_excel(request):

    if request.method == 'POST':

        archivo = request.FILES['archivo']
        workbook = load_workbook(archivo)
        hoja = workbook.active

        for fila in hoja.iter_rows(min_row=2):
            Alumno.objects.create(
                matricula=fila[0].value,
                nombre=fila[1].value,
                correo=fila[2].value,
                carrera=fila[3].value
            )

        messages.success(request, "Datos importados correctamente")
        return redirect('lista_alumnos')

    return render(request, 'alumnos/importar_excel.html')